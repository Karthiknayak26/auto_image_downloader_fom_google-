from openpyxl import load_workbook
from requests import get
from slugify import slugify
import os
import re
import time
import csv

# Your API credentials
api_key = "AIzaSyBS_O0vjR_28vWCY48XeZV86aX8hN6CFY0"
search_engine_id = "029fa609c5f34453a"

# Prepare images folder
if not os.path.exists("product_images"):
    os.makedirs("product_images")

# Load product names from the Excel file
wb = load_workbook("product_list_second api.xlsx")
ws = wb.active
product_names = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]

def clean_product_name(name):
    name = re.sub(r'\[.*?\]', '', name)  # Remove [bracketed] text
    name = re.sub(r'\d+ ?[Nn]?o?\.?', '', name)  # Remove '200 No'
    name = re.sub(r'\d+ ?[Gg]?m?\.?', '', name)  # Remove '50 Gm'
    name = re.sub(r'\W+', ' ', name)  # Remove special characters
    return ' '.join(word for word in name.split() if len(word) > 2).strip()

def get_image_url(query):
    url = "https://www.googleapis.com/customsearch/v1"
    params = {
        "key": api_key,
        "cx": search_engine_id,
        "q": query,
        "searchType": "image",
        "num": 1
    }
    try:
        response = get(url, params=params, timeout=10)
        data = response.json()
        if "items" in data:
            return data["items"][0]["link"]
        # Fallback: use cleaned query if different
        simple_query = clean_product_name(query)
        if simple_query and simple_query != query:
            params["q"] = simple_query
            response = get(url, params=params, timeout=10)
            data = response.json()
            if "items" in data:
                return data["items"][0]["link"]
    except Exception as e:
        print(f"Error fetching image for '{query}': {e}")
    return None

def download_image(image_url, file_name):
    try:
        response = get(image_url, stream=True, timeout=10)
        if response.status_code == 200:
            path = os.path.join("product_images", file_name)
            with open(path, "wb") as file:
                for chunk in response.iter_content(1024):
                    file.write(chunk)
            print(f"✅ Downloaded: {file_name}")
            return True
        else:
            print(f"❌ Failed to download: {file_name}")
    except Exception as e:
        print(f"Error downloading image {file_name}: {e}")
    return False

results = []
for product_name in product_names:
    slug_file_name = slugify(product_name) + ".jpg"
    image_url = get_image_url(product_name)
    if image_url:
        success = download_image(image_url, slug_file_name)
        status = "Downloaded" if success else "Failed to download"
    else:
        status = "No image found"
    results.append({
        "Product Name": product_name,
        "Status": status,
        "Image URL": image_url if image_url else ""
    })
    time.sleep(1)  # Respect API rate limits

# Save results to CSV
with open("image_download_results.csv", "w", newline="", encoding="utf-8") as csvfile:
    fieldnames = ["Product Name", "Status", "Image URL"]
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()
    for row in results:
        writer.writerow(row)
